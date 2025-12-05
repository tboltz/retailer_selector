# retail_selector/workbook.py
from __future__ import annotations

import asyncio
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional

import pandas as pd
import openpyxl

from .scraping import scrapingbee_fetch_many
from .parsing import hybrid_lookup_from_bee_result
from .logger import log


# --------------------------------------------------------------
# Load XLSX workbook safely into multiple pandas DataFrames
# --------------------------------------------------------------

def load_workbook_tables(workbook_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Loads all sheets from XLSX safely into pandas DataFrames.
    Returns dict: sheet_name -> DataFrame
    """
    log(f"Loading workbook: {workbook_path}", context="workbook")

    try:
        xl = pd.ExcelFile(workbook_path)
    except Exception as e:
        log(f"ERROR loading workbook: {e!r}", context="workbook")
        raise

    sheets: Dict[str, pd.DataFrame] = {}

    for name in xl.sheet_names:
        try:
            df = xl.parse(name)
            df.columns = [str(c).strip() for c in df.columns]
            sheets[name] = df
            log(f"Loaded sheet '{name}' with {len(df)} rows", context="workbook")
        except Exception as e:
            log(f"ERROR reading sheet {name}: {e!r}", context="workbook")

    return sheets


# --------------------------------------------------------------
# Extract Product↔Retailer Map from workbook
# --------------------------------------------------------------

def extract_product_map(sheets: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Returns the Product↔Retailer Map sheet as df.
    Auto-detects sheet by name.
    """
    possible_names = [
        "Product↔Retailer Map",
        "Product-Retailer Map",
        "Product Retailer Map",
        "Product Retailer",
        "Retailer Map",
    ]

    for name in possible_names:
        if name in sheets:
            df = sheets[name].copy()
            log("Product↔Retailer Map loaded from workbook.", context="workbook")
            return df

    raise KeyError("Could not find Product↔Retailer Map sheet in workbook.")


# --------------------------------------------------------------
# Save updated workbook back to disk
# --------------------------------------------------------------

def save_updated_workbook(
    workbook_path: Path,
    sheets: Dict[str, pd.DataFrame],
) -> Path:
    """
    Writes a fresh XLSX file with updated sheets.
    """
    log(f"Saving updated workbook → {workbook_path}", context="workbook")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, df in sheets.items():
        ws = wb.create_sheet(title=name)

        # Write header
        ws.append(list(df.columns))

        # Write rows
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))

    wb.save(workbook_path)
    log("Workbook saved.", context="workbook")
    return workbook_path


# --------------------------------------------------------------
# Main scanning logic for XLSX-driven workflow
# --------------------------------------------------------------

async def scan_workbook_async(
    workbook_path: Path,
    scrapingbee_api_key: str,
    limit: Optional[int] = None,
    concurrency: int = 20,
) -> Tuple[Path, pd.DataFrame]:
    """
    LOAD XLSX → extract Product↔Retailer Map → scrape → parse →
    update sheet → write XLSX → return path + updated df.
    """

    # ----------------------------------------------------------
    # Load workbook
    # ----------------------------------------------------------
    sheets = load_workbook_tables(workbook_path)
    df = extract_product_map(sheets)

    if df.empty:
        log("Product map is empty — nothing to scan.", context="workbook")
        sheets["Product↔Retailer Map"] = df
        save_updated_workbook(workbook_path, sheets)
        return workbook_path, df

    # ----------------------------------------------------------
    # Clean + filter URLs (same logic as orchestrator)
    # ----------------------------------------------------------
    df["search_url"] = df["search_url"].astype(str).str.strip()
    df = df[df["search_url"] != ""].copy()

    log(f"Filtered to {len(df)} rows with valid search_url", context="workbook")

    if df.empty:
        log("No valid URLs to scan.", context="workbook")
        sheets["Product↔Retailer Map"] = df
        save_updated_workbook(workbook_path, sheets)
        return workbook_path, df

    # Apply limit (like orchestrator)
    if limit is not None:
        df = df.head(limit).copy()
        log(f"Limit applied: scanning first {limit} valid rows", context="workbook")

    if df.empty:
        log("After limit, nothing to scan.", context="workbook")
        sheets["Product↔Retailer Map"] = df
        save_updated_workbook(workbook_path, sheets)
        return workbook_path, df

    # ----------------------------------------------------------
    # Ensure KPI columns
    # ----------------------------------------------------------
    kpis = [
        "In Stock (Y/N)",
        "Price ($USD)",
        "Last Scan (UTC)",
        "HTTP Status",
        "Parse Method",
        "Response ms",
        "Last Error",
        "URL Status",
        "Validation Issues",
    ]

    for col in kpis:
        if col not in df.columns:
            df[col] = None

    # ----------------------------------------------------------
    # Build lists for scanning
    # ----------------------------------------------------------
    urls = df["search_url"].tolist()
    df_indices = df.index.tolist()

    log(f"Fetching {len(urls)} URLs (concurrency={concurrency})", context="workbook")

    bee_results = await scrapingbee_fetch_many(
        urls=urls,
        api_key=scrapingbee_api_key,
        concurrency=concurrency,
    )

    now_iso = datetime.now(timezone.utc).isoformat()

    # ----------------------------------------------------------
    # Parse results & refill DataFrame
    # ----------------------------------------------------------
    for url, idx_in_df, bee in zip(urls, df_indices, bee_results):
        row = df.loc[idx_in_df]

        pid = str(row.get("product_id") or row.get("Product ID") or "").strip()
        desc = str(row.get("DESCRIPTION") or row.get("product_name") or "").strip()
        rkey = str(row.get("retailer_key") or row.get("Retailer") or "").strip()

        try:
            parsed = hybrid_lookup_from_bee_result(
                product_id=pid,
                description=desc,
                retailer_key=rkey,
                original_url=url,
                bee=bee,
            )

        except Exception as e:
            log(f"Parse exception row={idx_in_df} {e!r}", context="workbook")
            df.at[idx_in_df, "In Stock (Y/N)"] = ""
            df.at[idx_in_df, "Price ($USD)"] = float("nan")
            df.at[idx_in_df, "Last Error"] = f"parse_exception: {e!r}"
            df.at[idx_in_df, "URL Status"] = "error"
            df.at[idx_in_df, "HTTP Status"] = str(bee.get("status_code") or "")
            df.at[idx_in_df, "Response ms"] = float(bee.get("response_ms") or 0.0)
            df.at[idx_in_df, "Last Scan (UTC)"] = now_iso
            continue

        stock = parsed.get("stock")
        price = parsed.get("price")
        method = parsed.get("method")
        status = parsed.get("status")
        http_status = bee.get("status_code")
        elapsed_ms = parsed.get("response_ms")
        err = parsed.get("error")

        # Normalize stock
        if stock in (True, "Y", "y", "yes"):
            sf = "Y"
        elif stock in (False, "N", "n", "no"):
            sf = "N"
        else:
            sf = stock or ""

        df.at[idx_in_df, "In Stock (Y/N)"] = sf
        df.at[idx_in_df, "Price ($USD)"] = float(price) if price is not None else float("nan")
        df.at[idx_in_df, "Parse Method"] = method
        df.at[idx_in_df, "HTTP Status"] = str(http_status or "")
        df.at[idx_in_df, "URL Status"] = status
        df.at[idx_in_df, "Response ms"] = float(elapsed_ms or 0.0)
        df.at[idx_in_df, "Last Error"] = err or ""
        df.at[idx_in_df, "Last Scan (UTC)"] = now_iso

        log(
            f"row={idx_in_df} url={url} price={price} stock={sf} "
            f"method={method} status={status} http={http_status}",
            context="workbook",
        )

    # ----------------------------------------------------------
    # Update sheets + write back to disk
    # ----------------------------------------------------------
    sheets["Product↔Retailer Map"] = df

    save_updated_workbook(workbook_path, sheets)

    log("Workbook scan complete.", context="workbook")

    return workbook_path, df
